


CSV pandas


```python
# Leitura padrão BR
pd.read_csv('x.csv', sep=';', decimal=',', thousands='.', encoding='utf-8')

# Leitura com datas BR
pd.read_csv('x.csv', parse_dates=['data'], dayfirst=True)

# Leitura de parte
pd.read_csv('x.csv', usecols=['a', 'b'], nrows=100, skiprows=3)

# Força tipos
pd.read_csv('x.csv', dtype={'cpf': str, 'cep': str})

# Encoding alternativo
pd.read_csv('x.csv', encoding='latin-1')  # ou 'cp1252'

# Valores que viram NaN
pd.read_csv('x.csv', na_values=['-', 'n/a', 'NULL', '#N/D'])

# Escrita padrão BR (abre no Excel BR)
df.to_csv('y.csv', sep=';', decimal=',', index=False, encoding='utf-8-sig')

# Arquivo gigante em blocos
for chunk in pd.read_csv('grande.csv', chunksize=50000):
    processar(chunk)
```
CSV módulo nativo
```python
import csv

# Ler como dict
with open('x.csv', encoding='utf-8', newline='') as f:
    for linha in csv.DictReader(f, delimiter=';'):
        print(linha['nome'])

# Escrever como dict
with open('y.csv', 'w', encoding='utf-8', newline='') as f:
    escritor = csv.DictWriter(f, fieldnames=['nome', 'idade'], delimiter=';')
    escritor.writeheader()
    escritor.writerows([{'nome': 'Ana', 'idade': 30}])

# Detectar delimitador automaticamente
with open('x.csv') as f:
    dialeto = csv.Sniffer().sniff(f.read(4096))
    print(dialeto.delimiter)
```
Excel pandas
```python
# Leitura
pd.read_excel('x.xlsx', sheet_name='Vendas')   # aba pelo nome
pd.read_excel('x.xlsx', sheet_name=0)          # aba pelo índice
pd.read_excel('x.xlsx', sheet_name=None)       # todas (dict)
pd.read_excel('x.xlsx', sheet_name=['A','B'])  # lista

# Listar abas sem ler
pd.ExcelFile('x.xlsx').sheet_names

# Pular linhas iniciais
pd.read_excel('x.xlsx', skiprows=3)

# Range específico
pd.read_excel('x.xlsx', usecols='B:F', skiprows=2, nrows=50)

# Escrita
df.to_excel('y.xlsx', index=False, sheet_name='Resultado')

# Várias abas
with pd.ExcelWriter('y.xlsx') as writer:
    df1.to_excel(writer, sheet_name='Aba1', index=False)
    df2.to_excel(writer, sheet_name='Aba2', index=False)

# Adicionar aba sem apagar outras
with pd.ExcelWriter('x.xlsx', mode='a', if_sheet_exists='replace') as w:
    df_novo.to_excel(w, sheet_name='Nova')
```
Excel openpyxl (formatação)
```python
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Abrir
wb = load_workbook('x.xlsx')
ws = wb.active              # aba ativa
ws = wb['Nome']             # por nome
ws = wb.worksheets[0]       # por índice

# Ler célula
valor = ws['A1'].value
valor = ws.cell(row=1, column=1).value

# Escrever
ws['A1'] = 'Total'
ws.cell(row=1, column=1, value='Total')

# Iterar
for linha in ws.iter_rows(min_row=2, values_only=True):
    print(linha)

# Formatação
ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
ws['A1'].fill = PatternFill('solid', fgColor='4472C4')
ws['A1'].alignment = Alignment(horizontal='center')
ws['A1'].number_format = 'R$ #,##0.00'

# Bordas
borda = Border(left=Side('thin'), right=Side('thin'),
               top=Side('thin'), bottom=Side('thin'))
ws['A1'].border = borda

# Largura de coluna
ws.column_dimensions['A'].width = 20

# Congelar painel (mantém cabeçalho fixo)
ws.freeze_panes = 'A2'

# Filtro automático
ws.auto_filter.ref = ws.dimensions

# Mesclar
ws.merge_cells('A1:D1')

# Fórmula
ws['D2'] = '=B2*C2'
ws['D10'] = '=SUM(D2:D9)'

# Salvar
wb.save('x.xlsx')
```
Formatos numéricos úteis
```
'#,##0'              -> 1.234
'#,##0.00'           -> 1.234,56
'R$ #,##0.00'        -> R$ 1.234,56
'0.00%'              -> 12,34%
'0.00E+00'           -> 1,23E+03 (notação científica)
'dd/mm/yyyy'         -> 15/08/2025
'dd/mm/yyyy hh:mm'   -> 15/08/2025 14:30
'mmm-yy'             -> ago-25
'#,##0;[Red]-#,##0'  -> negativos em vermelho
```
Conversões rápidas
```python
# CSV para Excel
pd.read_csv('x.csv').to_excel('x.xlsx', index=False)

# Excel para CSV BR
pd.read_excel('x.xlsx').to_csv('x.csv', sep=';', decimal=',',
                                 index=False, encoding='utf-8-sig')

# Várias abas do Excel em CSVs separados
abas = pd.read_excel('x.xlsx', sheet_name=None)
for nome, df in abas.items():
    df.to_csv(f'{nome}.csv', sep=';', decimal=',', index=False,
              encoding='utf-8-sig')

# Juntar vários CSVs
import glob
dfs = [pd.read_csv(f) for f in glob.glob('*.csv')]
pd.concat(dfs, ignore_index=True).to_csv('total.csv', index=False)
```
Armadilhas comuns
Encoding. Se acentos aparecem como  , troque de utf-8 para latin-1 ou cp1252. Para CSV abrir direto no Excel BR, use `encoding='utf-8-sig'`.
Separador. Excel brasileiro usa ponto-e-vírgula, não vírgula. Sempre teste o separador.
Decimal. Brasil usa vírgula. Se você salva sem `decimal=','`, o Excel BR interpreta 1.234 como 1234 (mil duzentos e trinta e quatro) e não 1,234.
Data. Sem `dayfirst=True`, 03/04/2025 vira 4 de março em vez de 3 de abril.
Dtype. CPF e CEP começando com zero somem se o pandas ler como int. Force `dtype={'cpf': str}`.
sheet_name=None. Retorna DICT, não DataFrame. Sempre confira antes de usar.
to_excel com index=True. Se esquecer `index=False`, a primeira coluna do Excel vira o índice do DataFrame, geralmente indesejado.
openpyxl data_only. Se você abre um Excel criado por código e quer os valores calculados das fórmulas, precisa `data_only=True`. Mas isso só funciona se o Excel foi aberto no aplicativo antes, porque o openpyxl não calcula fórmulas.
Arquivos grandes. Excel tem limite de ~1 milhão de linhas por aba. Para volumes maiores, use CSV ou Parquet.