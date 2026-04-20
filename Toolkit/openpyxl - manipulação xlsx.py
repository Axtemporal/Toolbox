"""
excel_openpyxl.py
Controle fino de Excel com openpyxl.
Use para formatação, fórmulas, cores, células mescladas.
Para fluxo padrão de ler/escrever tabela, prefira excel_pandas.py.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule


# ============================================================
# Abrir e salvar
# ============================================================

def novo_workbook():
    """
    Cria um Workbook vazio em memória.

    Exemplo:
        wb = novo_workbook()
        ws = wb.active
        ws['A1'] = 'Olá'
        wb.save('arquivo.xlsx')
    """
    return Workbook()


def abrir_excel(caminho, somente_leitura=False):
    """
    Abre Excel existente. somente_leitura=True é muito mais rápido
    para arquivos grandes quando você só quer ler.
    """
    return load_workbook(caminho, read_only=somente_leitura, data_only=False)


def abrir_calculando_formulas(caminho):
    """
    Abre Excel com os valores calculados das fórmulas.
    data_only=True faz openpyxl retornar o valor em cache em vez da fórmula.
    Requer que o Excel tenha sido aberto e salvo antes pelo Excel (não pelo openpyxl).
    """
    return load_workbook(caminho, data_only=True)


# ============================================================
# Navegar celulas e abas
# ============================================================

def listar_abas_wb(wb):
    """Lista nomes das abas do Workbook."""
    return wb.sheetnames


def obter_aba(wb, nome_ou_indice=0):
    """
    Retorna aba pelo nome ou índice.
    """
    if isinstance(nome_ou_indice, int):
        return wb.worksheets[nome_ou_indice]
    return wb[nome_ou_indice]


def adicionar_aba(wb, nome, posicao=None):
    """Cria nova aba. posicao=0 coloca como primeira."""
    ws = wb.create_sheet(title=nome, index=posicao)
    return ws


def remover_aba(wb, nome):
    """Remove aba pelo nome."""
    del wb[nome]


def ler_celula(ws, endereco):
    """
    Lê valor de uma célula.

    Exemplo:
        valor = ler_celula(ws, 'A1')
    """
    return ws[endereco].value


def escrever_celula(ws, endereco, valor):
    """Escreve valor em célula."""
    ws[endereco] = valor


def ler_intervalo(ws, range_str):
    """
    Lê intervalo como lista de listas.

    Exemplo:
        valores = ler_intervalo(ws, 'A1:C10')
    """
    return [[c.value for c in linha] for linha in ws[range_str]]


def escrever_intervalo(ws, inicio, matriz):
    """
    Escreve matriz (lista de listas) a partir de uma célula.

    Exemplo:
        escrever_intervalo(ws, 'A1', [
            ['Nome', 'Idade'],
            ['Ana', 30],
            ['Bruno', 25]
        ])
    """
    col_ini = column_index_from_string(''.join(c for c in inicio if c.isalpha()))
    linha_ini = int(''.join(c for c in inicio if c.isdigit()))

    for i, linha in enumerate(matriz):
        for j, valor in enumerate(linha):
            ws.cell(row=linha_ini + i, column=col_ini + j, value=valor)


def dimensoes(ws):
    """
    Retorna dict com número de linhas e colunas com dados.
    """
    return {
        'max_linha': ws.max_row,
        'max_coluna': ws.max_column,
        'range': ws.dimensions  # tipo 'A1:D10'
    }


# ============================================================
# Iterar sobre dados
# ============================================================

def iterar_linhas(ws, min_linha=1, max_linha=None, com_valores=True):
    """
    Gera tuplas com valores de cada linha.
    Mais eficiente que ler tudo de uma vez.

    Exemplo:
        for linha in iterar_linhas(ws, min_linha=2):  # pula cabeçalho
            print(linha)
    """
    for linha in ws.iter_rows(min_row=min_linha, max_row=max_linha,
                               values_only=com_valores):
        yield linha


def iterar_colunas(ws, min_col=1, max_col=None, com_valores=True):
    """Gera tuplas com valores de cada coluna."""
    for col in ws.iter_cols(min_col=min_col, max_col=max_col,
                             values_only=com_valores):
        yield col


def ws_para_dicts(ws, linha_cabecalho=1):
    """
    Converte aba em lista de dicts usando uma linha como cabeçalho.

    Exemplo:
        registros = ws_para_dicts(ws)
    """
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        return []
    cabecalho = linhas[linha_cabecalho - 1]
    dados = linhas[linha_cabecalho:]
    return [dict(zip(cabecalho, linha)) for linha in dados]


# ============================================================
# Formatacao
# ============================================================

def formatar_negrito(ws, range_str):
    """
    Aplica negrito num intervalo.

    Exemplo:
        formatar_negrito(ws, 'A1:E1')  # cabeçalho em negrito
    """
    for linha in ws[range_str]:
        for celula in linha:
            celula.font = Font(bold=True)


def formatar_cabecalho(ws, linha=1, cor_fundo='4472C4', cor_fonte='FFFFFF'):
    """
    Formatação clássica de cabeçalho: fundo azul, texto branco, negrito.
    """
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        celula = ws.cell(row=linha, column=col)
        celula.font = Font(bold=True, color=cor_fonte)
        celula.fill = PatternFill('solid', fgColor=cor_fundo)
        celula.alignment = Alignment(horizontal='center', vertical='center')


def aplicar_cor_fundo(ws, range_str, cor_hex):
    """
    Pinta fundo de células. cor_hex sem o #.

    Exemplo:
        aplicar_cor_fundo(ws, 'B2:B10', 'FFE699')  # amarelo claro
    """
    fill = PatternFill('solid', fgColor=cor_hex)
    for linha in ws[range_str]:
        for celula in linha:
            celula.fill = fill


def aplicar_bordas(ws, range_str, estilo='thin'):
    """
    Adiciona bordas ao redor de cada célula do intervalo.
    estilo: 'thin', 'medium', 'thick', 'dashed'
    """
    borda = Border(
        left=Side(style=estilo),
        right=Side(style=estilo),
        top=Side(style=estilo),
        bottom=Side(style=estilo)
    )
    for linha in ws[range_str]:
        for celula in linha:
            celula.border = borda


def formato_numero(ws, range_str, formato):
    """
    Aplica formato numérico a células.

    Formatos comuns:
        '#,##0.00'             # número com 2 decimais
        '#,##0'                # inteiro com separador
        'R$ #,##0.00'          # moeda BR
        '0.00%'                # percentual
        'dd/mm/yyyy'           # data BR
        'dd/mm/yyyy hh:mm'     # data e hora
        '#,##0.00;[Red]-#,##0.00'  # negativos em vermelho

    Exemplo:
        formato_numero(ws, 'B2:B100', 'R$ #,##0.00')
    """
    for linha in ws[range_str]:
        for celula in linha:
            celula.number_format = formato


def ajustar_largura_colunas(ws, larguras):
    """
    Define larguras de colunas.

    Exemplo:
        ajustar_largura_colunas(ws, {'A': 20, 'B': 15, 'C': 12})
    """
    for letra, largura in larguras.items():
        ws.column_dimensions[letra].width = largura


def auto_ajustar_largura(ws):
    """
    Ajusta largura automaticamente baseado no conteúdo.
    Não perfeito, mas resolve na maioria dos casos.
    """
    for col in ws.columns:
        max_tamanho = 0
        letra = get_column_letter(col[0].column)
        for celula in col:
            try:
                tam = len(str(celula.value)) if celula.value is not None else 0
                max_tamanho = max(max_tamanho, tam)
            except Exception:
                pass
        ws.column_dimensions[letra].width = min(max_tamanho + 2, 50)


# ============================================================
# Painel congelado e filtros
# ============================================================

def congelar_painel(ws, celula='A2'):
    """
    Congela painel acima e à esquerda da célula informada.
    'A2' congela só a primeira linha. 'B2' congela primeira linha E primeira coluna.
    """
    ws.freeze_panes = celula


def adicionar_autofiltro(ws, range_str=None):
    """
    Ativa filtros automáticos na aba.
    Se range_str=None, usa todo o range com dados.
    """
    ws.auto_filter.ref = range_str or ws.dimensions


# ============================================================
# Formulas
# ============================================================

def inserir_formula(ws, endereco, formula):
    """
    Insere fórmula numa célula.
    A fórmula DEVE começar com '='.

    Exemplo:
        inserir_formula(ws, 'D2', '=B2*C2')
        inserir_formula(ws, 'D10', '=SUM(D2:D9)')
    """
    ws[endereco] = formula


def somar_coluna(ws, coluna, linha_inicio, linha_fim, linha_total):
    """
    Insere fórmula de soma numa célula.

    Exemplo:
        somar_coluna(ws, 'D', 2, 100, 101)  # =SUM(D2:D100) na célula D101
    """
    ws[f'{coluna}{linha_total}'] = f'=SUM({coluna}{linha_inicio}:{coluna}{linha_fim})'


# ============================================================
# Formatacao condicional
# ============================================================

def escala_de_cor(ws, range_str, cor_baixo='FFFFFF', cor_alto='63BE7B'):
    """
    Aplica escala de cor (mapa de calor).
    Cores padrão: branco para o menor, verde para o maior.
    """
    regra = ColorScaleRule(
        start_type='min', start_color=cor_baixo,
        end_type='max', end_color=cor_alto
    )
    ws.conditional_formatting.add(range_str, regra)


def destacar_maiores_que(ws, range_str, limite, cor_hex='FF9999'):
    """
    Destaca células com valor maior que limite.

    Exemplo:
        destacar_maiores_que(ws, 'D2:D100', 1000)  # vermelho se > 1000
    """
    regra = CellIsRule(
        operator='greaterThan',
        formula=[str(limite)],
        fill=PatternFill('solid', fgColor=cor_hex)
    )
    ws.conditional_formatting.add(range_str, regra)


# ============================================================
# Mesclar celulas
# ============================================================

def mesclar(ws, range_str):
    """
    Mescla células num intervalo.

    Exemplo:
        mesclar(ws, 'A1:E1')  # mescla primeira linha
    """
    ws.merge_cells(range_str)


def desmesclar(ws, range_str):
    """Desfaz a mesclagem."""
    ws.unmerge_cells(range_str)


if __name__ == '__main__':
    # Teste rápido
    wb = novo_workbook()
    ws = wb.active
    ws.title = 'Teste'
    escrever_intervalo(ws, 'A1', [
        ['Produto', 'Preço', 'Quantidade'],
        ['Café', 25.50, 100],
        ['Açúcar', 5.20, 50]
    ])
    formatar_cabecalho(ws)
    formato_numero(ws, 'B2:B3', 'R$ #,##0.00')
    auto_ajustar_largura(ws)
    congelar_painel(ws)
    wb.save('/tmp/teste_openpyxl.xlsx')
    print('Arquivo criado em /tmp/teste_openpyxl.xlsx')
